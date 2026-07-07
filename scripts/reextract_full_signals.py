"""
reextract_full_signals.py — re-segment the FULL cleaned signals ourselves (multi-burst
activity detection) instead of using the lab's pre-segments, then re-run the decisive test.

Honest long-shot: the binding constraint is label scarcity (5 VBmax/tool, 18 tools, no
replication), not feature richness. This checks whether our own activity cuts + features carry
generalizable per-tool signal that the lab's segments did not. Reads E:\...\Cleaned with Time.
Writes data/input/derived/features_reextracted.csv and prints the LOO-CV verdict.
"""
import os, re, sys
import numpy as np
import pandas as pd

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, os.path.join(ROOT, "src"))
from phm.segmentation_p8 import load_signal, time_features, freq_features, wavelet_features, FS
from phm.feature_selection_p8 import select_topk, shap_scores

CLEAN = r"E:\Hanwha\Hamed Thesis\NEW DOE\Vibration\Cleaned with Time"
EXCEL = r"C:\Users\Administrador\Desktop\NEW DOE.xlsx"


def detect_bursts(v, fs=FS, min_s=0.05, merge_s=0.05):
    win = max(1, int(0.005 * fs))
    env = np.convolve(np.abs(v), np.ones(win) / win, mode="same")
    base = np.percentile(env, 10); thr = base + 0.2 * (env.max() - base)
    active = env > thr
    if not active.any():
        return [(0, len(v) - 1)]
    idx = np.flatnonzero(active)
    cuts = np.where(np.diff(idx) > int(merge_s * fs))[0]
    groups = np.split(idx, cuts + 1)
    out = [(int(g[0]), int(g[-1])) for g in groups if (g[-1] - g[0]) > int(min_s * fs)]
    return out or [(int(idx[0]), int(idx[-1]))]


def _read_v(path):
    df = pd.read_csv(path)
    col = df.iloc[:, 0] if df.shape[1] == 1 else df.iloc[:, 1]   # 1-col=Value; 2-col=Time,Value
    v = pd.to_numeric(col, errors="coerce").to_numpy(float)
    v = v[~np.isnan(v)]
    return v - np.nanmean(v)


def file_features(path):
    v = _read_v(path)
    feats = []
    for s, e in detect_bursts(v):
        w = v[s:e + 1]
        if len(w) < 8:
            continue
        d = {}; d.update(time_features(w)); d.update(freq_features(w, FS)); d.update(wavelet_features(w))
        feats.append(d)
    if not feats:
        return {}
    fdf = pd.DataFrame(feats)
    out = {}
    for c in fdf.columns:
        out[f"{c}__mean"] = float(fdf[c].mean()); out[f"{c}__std"] = float(fdf[c].std())
    return out


def excel_map():
    import openpyxl
    ws = openpyxl.load_workbook(EXCEL, data_only=True)["DOE"]
    mp = {}
    for r in range(2, ws.max_row + 1):
        wp, tool, vb = ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 17).value
        if wp is None or tool is None:
            continue
        mp[int(wp)] = (int(tool), vb)
    return mp


def main():
    mp = excel_map()
    files = {}
    for f in os.listdir(CLEAN):
        m = re.match(r"^([AR])(\d+)_", f)
        if m:
            files.setdefault(int(m.group(2)), {})[m.group(1).upper()] = os.path.join(CLEAN, f)
    rows = []
    done = 0
    for n in sorted(files):
        if n not in mp:
            continue
        tool, vb = mp[n]
        ch = files[n]
        af = file_features(ch["A"]) if "A" in ch else {}
        rf = file_features(ch["R"]) if "R" in ch else {}
        if not af and not rf:
            continue
        rec = {"experiment_id": n, "tool_id": f"T{tool}", "vb_um": vb}
        rec.update({f"A_{k}": v for k, v in af.items()})
        rec.update({f"R_{k}": v for k, v in rf.items()})
        rows.append(rec); done += 1
        if done % 20 == 0:
            print(f"  ...processed {done} experiments", flush=True)
    df = pd.DataFrame(rows)
    out = os.path.join(ROOT, "data", "input", "derived", "features_reextracted.csv")
    df.to_csv(out, index=False)
    print(f"wrote {out}: {df.shape[0]} exp x {df.shape[1]} cols", flush=True)

    # decisive LOO-CV test on the re-extracted features (per-tool wear rate)
    from sklearn.linear_model import RidgeCV
    from sklearn.model_selection import LeaveOneOut
    feat = [c for c in df.columns if c not in ("experiment_id", "tool_id", "vb_um")]
    tdat = []
    for t, g in df.groupby("tool_id"):
        g = g.sort_values("experiment_id")
        if g.vb_um.notna().sum() < 2:
            continue
        rate = np.polyfit(np.arange(len(g)), g.vb_um.fillna(method="ffill").values.astype(float), 1)[0]
        rec = {"tool": t, "rate": rate}; rec.update(g[feat].mean().to_dict()); tdat.append(rec)
    D = pd.DataFrame(tdat).fillna(0.0); y = D["rate"].values; X = D[feat].fillna(0.0)
    pred = np.zeros(len(y))
    for tr, te in LeaveOneOut().split(X):
        sel, _ = select_topk(X.iloc[tr], y[tr], feat, k=8)
        mu, sd = X.iloc[tr][sel].mean(), X.iloc[tr][sel].std() + 1e-9
        m = RidgeCV(alphas=[1, 10, 100, 1000]).fit((X.iloc[tr][sel] - mu) / sd, y[tr])
        pred[te] = m.predict((X.iloc[te][sel] - mu) / sd)
    r2 = 1 - np.sum((y - pred) ** 2) / np.sum((y - y.mean()) ** 2)
    print(f"\nRE-EXTRACTED features -> per-tool rate LOO-CV R2 = {r2:+.2f}  (>0 = beat the null!)", flush=True)
    print("VERDICT: " + ("RE-EXTRACTION FOUND SIGNAL (R2>0)!" if r2 > 0.05
                         else "still null (R2<=0) -> confirms label scarcity, not feature quality, is the limit."), flush=True)


if __name__ == "__main__":
    main()
