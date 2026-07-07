"""
extract_features_from_disk.py — stream multi-domain features off the external disk.

Source of truth = the lab's SEGMENTED vibration files (small): {A,R}<n>_p<k>.txt, each a
"Time,Value" contact window at 50 kHz. The heavy "Cleaned with Time" full files are NOT read.
Tool / VBmax / cutting conditions come from the DOE Excel (experiment number = join key).
Reuses the official P8 feature set (src/phm/segmentation_p8.py) for consistency with the
frozen T14 (=old "T01") build. Writes only a small per-experiment features table + a
multi-tool manifest; the heavy raw data stays on the disk.

Usage:
  # validation on a few experiments:
  python scripts/extract_features_from_disk.py --exps 1 66 71 94 --validate
  # full run:
  python scripts/extract_features_from_disk.py
"""
import argparse, os, re, sys
import numpy as np
import pandas as pd

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, os.path.join(ROOT, "src"))
from phm.segmentation_p8 import load_signal, detect_contact, segment_features

SEG_DEFAULT  = r"E:\Hanwha\Hamed Thesis\NEW DOE\Vibration\Segmented"
EXCEL_DEFAULT = r"C:\Users\Administrador\Desktop\NEW DOE.xlsx"
OUT_DIR = os.path.join(ROOT, "data", "input", "derived")
MIN_SEG_RELIABLE = 5            # <5 valid segments -> reliability flag (like exp77)
AGG = ("mean", "std", "median")


def _resolve_vbmax(ws_val, ws_fml, col=17):
    """Robust VBmax per row: literal if present, else evaluate the user's =AVERAGE(...)
    interpolation formulas recursively (their cache may be stale after a programmatic save)."""
    cache = {}

    def val(r):
        if r in cache:
            return cache[r]
        cache[r] = None  # guard against cycles
        v = ws_val.cell(r, col).value
        if isinstance(v, (int, float)):
            cache[r] = float(v); return cache[r]
        f = ws_fml.cell(r, col).value
        if isinstance(f, str) and f.upper().startswith("=AVERAGE"):
            refs = re.findall(r"[A-Z]+(\d+)", f)             # Q117, Q120 -> rows
            xs = [val(int(rr)) for rr in refs]
            xs = [x for x in xs if x is not None]
            cache[r] = float(np.mean(xs)) if xs else None
        return cache[r]
    return val


def read_excel_map(path):
    """experiment number -> dict(tool, order, vbmax, vc, fz, cool, strat)."""
    import openpyxl
    ws_val = openpyxl.load_workbook(path, data_only=True)["DOE"]
    ws_fml = openpyxl.load_workbook(path, data_only=False)["DOE"]
    getvb = _resolve_vbmax(ws_val, ws_fml)
    ws = ws_val
    rows = []
    for r in range(2, ws.max_row + 1):
        wp = ws.cell(r, 2).value
        tool = ws.cell(r, 3).value
        if wp is None or tool is None:
            continue
        rows.append(dict(exp=int(wp), tool=int(tool), vbmax=getvb(r),
                         vc=ws.cell(r, 7).value, fz=ws.cell(r, 8).value,
                         cool=str(ws.cell(r, 6).value), strat=str(ws.cell(r, 4).value)))
    df = pd.DataFrame(rows).sort_values(["tool", "exp"]).reset_index(drop=True)
    df["order"] = df.groupby("tool").cumcount() + 1     # within-tool order
    return {row.exp: row for row in df.itertuples()}


def discover(seg_dir):
    """experiment number -> {'A':[paths], 'R':[paths]} from segment filenames."""
    exps = {}
    for f in os.listdir(seg_dir):
        m = re.match(r"^([AR])(\d+)_p(\d+)\.txt$", f, re.I)
        if not m:
            continue
        ch, n = m.group(1).upper(), int(m.group(2))
        exps.setdefault(n, {"A": [], "R": []})[ch].append(os.path.join(seg_dir, f))
    return exps


def channel_features(paths):
    """Aggregate per-segment features (mean/std/median over segments) for one channel."""
    per_seg = []
    for p in sorted(paths):
        try:
            _, v = load_signal(p)
            if len(v) < 8:
                continue
            c = detect_contact(v)
            per_seg.append(segment_features(v, c["start"], c["end"]))
        except Exception as e:
            print(f"    WARN {os.path.basename(p)}: {e}")
    if not per_seg:
        return {}, 0
    fdf = pd.DataFrame(per_seg)
    base = [c for c in fdf.columns if c not in ("segment_n_samples", "segment_duration_s")]
    out = {}
    for col in base:
        s = fdf[col].astype(float)
        out[f"{col}__mean"] = float(s.mean())
        out[f"{col}__std"] = float(s.std())
        out[f"{col}__median"] = float(s.median())
    return out, len(per_seg)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--seg", default=SEG_DEFAULT)
    ap.add_argument("--excel", default=EXCEL_DEFAULT)
    ap.add_argument("--exps", type=int, nargs="*", help="subset of experiment numbers")
    ap.add_argument("--validate", action="store_true", help="print rows, do not overwrite full outputs")
    args = ap.parse_args()

    emap = read_excel_map(args.excel)
    found = discover(args.seg)
    targets = args.exps if args.exps else sorted(emap)      # only experiments in the Excel
    feat_rows, man_rows = [], []
    for n in targets:
        meta = emap.get(n)
        if meta is None:
            print(f"exp {n}: not in Excel -> skip (no tool/VBmax)")
            continue
        ch = found.get(n, {"A": [], "R": []})
        af, na = channel_features(ch["A"])
        rf, nr = channel_features(ch["R"])
        has_signal = (na > 0 and nr > 0)
        row = {"experiment_id": n, "tool_id": f"T{meta.tool}", "within_tool_order": meta.order,
               "vb_um": meta.vbmax, "vc": meta.vc, "fz": meta.fz, "cooling": meta.cool}
        row.update({f"A_{k}": v for k, v in af.items()})
        row.update({f"R_{k}": v for k, v in rf.items()})
        feat_rows.append(row)
        man_rows.append({"tool_id": f"T{meta.tool}", "experiment_id": n,
                         "within_tool_order": meta.order, "vb_um": meta.vbmax,
                         "has_signal": has_signal, "n_seg_A": na, "n_seg_R": nr,
                         "usable_sensor": bool(has_signal and min(na, nr) >= 1),
                         "usable_trajectory": meta.vbmax is not None,
                         "reliability_flag": "" if (na >= MIN_SEG_RELIABLE and nr >= MIN_SEG_RELIABLE)
                         else ("signal_less" if not has_signal else "reduced_segments"),
                         "vc": meta.vc, "fz": meta.fz, "cooling": meta.cool})
        print(f"exp {n:>3} T{meta.tool:<2} ord{meta.order:<2} VB={meta.vbmax} "
              f"segA={na} segR={nr} feats={len(af)+len(rf)} "
              f"{'' if has_signal else '<<signal-less'}")

    fdf = pd.DataFrame(feat_rows)
    mdf = pd.DataFrame(man_rows)
    if args.validate:
        print(f"\nVALIDATION: {len(fdf)} experiments, {fdf.shape[1]} columns "
              f"({fdf.shape[1]-7} features). usable_sensor={mdf['usable_sensor'].sum()} "
              f"trajectory={mdf['usable_trajectory'].sum()}")
        return
    os.makedirs(OUT_DIR, exist_ok=True)
    fdf.to_csv(os.path.join(OUT_DIR, "features_experiment.csv"), index=False)
    mdf.to_csv(os.path.join(OUT_DIR, "manifest_multitool.csv"), index=False)
    print(f"\nwrote {OUT_DIR}/features_experiment.csv  ({fdf.shape[0]} rows x {fdf.shape[1]} cols)")
    print(f"wrote {OUT_DIR}/manifest_multitool.csv")


if __name__ == "__main__":
    main()
